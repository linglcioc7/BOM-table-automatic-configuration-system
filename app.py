from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
import tempfile
import uuid
import pytz
from functools import wraps

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///bom_system.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'bom_system_secret_key_2025'  # 用于session加密
app.config['SESSION_COOKIE_SECURE'] = False  # 开发环境设为False
app.config['SESSION_COOKIE_HTTPONLY'] = True  # 防止XSS攻击
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # 跨站请求保护
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 会话有效期24小时
db = SQLAlchemy(app)

# 登录验证装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# 上海时区函数
def get_shanghai_time():
    """获取上海时区的当前时间"""
    shanghai_tz = pytz.timezone('Asia/Shanghai')
    return datetime.now(shanghai_tz)

# 数据模型
class Recipe(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    recipe_name = db.Column(db.String(100), nullable=False, unique=True)  # 配方名称唯一性约束
    description = db.Column(db.Text)
    product_category = db.Column(db.String(100))  # 产品类别，可为空
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=get_shanghai_time)
    updated_at = db.Column(db.DateTime, default=get_shanghai_time, onupdate=get_shanghai_time)

class RecipeItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    recipe_id = db.Column(db.Integer, db.ForeignKey('recipe.id'), nullable=False)
    material_code = db.Column(db.String(50), nullable=False)
    material_name = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(20), nullable=False)
    line_number = db.Column(db.String(10), nullable=False)  # 改为字符串格式
    project_category = db.Column(db.String(10), default='L')

class BOMRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    parent_material_code = db.Column(db.String(50), nullable=False)
    parent_material_name = db.Column(db.String(200), nullable=False)
    basic_quantity = db.Column(db.Float, nullable=False)
    basic_unit = db.Column(db.String(20), nullable=False)
    recipe_ids = db.Column(db.Text, nullable=False)  # 存储多个配方ID，用逗号分隔
    created_at = db.Column(db.DateTime, default=get_shanghai_time)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == 'shwx' and password == 'shwxjsb':
            session.permanent = True  # 设置永久会话
            session['logged_in'] = True
            session['username'] = username
            session['login_time'] = datetime.now().isoformat()
            return redirect(url_for('admin'))
        else:
            return render_template('admin_login.html', error='用户名或密码错误')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    return redirect(url_for('admin_login'))

@app.route('/admin')
@login_required
def admin():
    return render_template('admin.html')

@app.route('/api/recipes')
def get_recipes():
    recipes = Recipe.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': r.id,
        'name': r.recipe_name,
        'description': r.description,
        'product_category': r.product_category or '',  # 添加产品类别
        'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
        'updated_at': r.updated_at.strftime('%Y-%m-%d %H:%M:%S') if r.updated_at else ''
    } for r in recipes])

@app.route('/api/session/status')
def check_session_status():
    """检查用户会话状态"""
    return jsonify({
        'logged_in': session.get('logged_in', False),
        'username': session.get('username', ''),
        'login_time': session.get('login_time', '')
    })

@app.route('/api/recipe/categories')
def get_recipe_categories():
    """获取所有现有的产品类别"""
    categories = db.session.query(Recipe.product_category).filter(
        Recipe.is_active == True,
        Recipe.product_category.isnot(None),
        Recipe.product_category != ''
    ).distinct().all()
    
    # 提取类别名称并排序
    category_list = [cat[0] for cat in categories if cat[0]]
    category_list.sort()
    
    # 添加"未分类"选项
    category_list.append("未分类")
    
    return jsonify(category_list)

@app.route('/api/recipe/<int:recipe_id>')
@login_required
def get_recipe_items(recipe_id):
    items = RecipeItem.query.filter_by(recipe_id=recipe_id).order_by(RecipeItem.line_number).all()
    return jsonify([{
        'id': item.id,
        'material_code': item.material_code,
        'material_name': item.material_name,
        'quantity': item.quantity,
        'unit': item.unit,
        'line_number': item.line_number,
        'project_category': item.project_category
    } for item in items])

@app.route('/api/recipe/<int:recipe_id>', methods=['PUT'])
@login_required
def update_recipe(recipe_id):
    data = request.json
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({'success': False, 'message': '配方不存在'}), 404
    
    recipe.recipe_name = data['name']
    recipe.description = data['description']
    recipe.product_category = data.get('product_category', '')  # 更新产品类别
    recipe.updated_at = get_shanghai_time()
    
    # 删除旧的配方项
    RecipeItem.query.filter_by(recipe_id=recipe_id).delete()
    
    # 添加新的配方项
    for item in data['items']:
        recipe_item = RecipeItem(
            recipe_id=recipe.id,
            material_code=item['material_code'],
            material_name=item['material_name'],
            quantity=item['quantity'],
            unit=item['unit'],
            line_number=item['line_number'],
            project_category=item['project_category']
        )
        db.session.add(recipe_item)
    
    db.session.commit()
    return jsonify({'success': True, 'message': '配方更新成功'})

@app.route('/api/recipe/<int:recipe_id>', methods=['DELETE'])
@login_required
def delete_recipe(recipe_id):
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({'success': False, 'message': '配方不存在'}), 404
    
    # 软删除：设置is_active为False
    recipe.is_active = False
    db.session.commit()
    
    return jsonify({'success': True, 'message': '配方删除成功'})

@app.route('/api/generate_bom', methods=['POST'])
def generate_bom():
    data = request.json
    
    # 检查必要字段
    if not all(key in data for key in ['parent_material_code', 'parent_material_name', 'basic_quantity', 'basic_unit', 'recipe_ids']):
        return jsonify({'success': False, 'message': '缺少必要字段'}), 400
    
    if not data['recipe_ids']:
        return jsonify({'success': False, 'message': '请选择至少一个配方'}), 400
    
    # 创建BOM请求记录
    bom_request = BOMRequest(
        parent_material_code=data['parent_material_code'],
        parent_material_name=data['parent_material_name'],
        basic_quantity=data['basic_quantity'],
        basic_unit=data['basic_unit'],
        recipe_ids=','.join(map(str, data['recipe_ids']))  # 将配方ID数组转换为逗号分隔的字符串
    )
    db.session.add(bom_request)
    db.session.commit()
    
    # 获取所有选中的配方信息
    all_items = []
    for recipe_id in data['recipe_ids']:
        recipe = Recipe.query.get(recipe_id)
        if recipe:
            items = RecipeItem.query.filter_by(recipe_id=recipe_id).order_by(RecipeItem.line_number).all()
            # 为每个配方项添加配方信息
            for item in items:
                all_items.append({
                    'recipe_name': recipe.recipe_name,
                    'recipe_description': recipe.description,
                    'line_number': item.line_number,
                    'material_code': item.material_code,
                    'material_name': item.material_name,
                    'quantity': item.quantity,
                    'unit': item.unit,
                    'project_category': item.project_category
                })
    
    # 生成Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM表"
    
    # 设置列宽
    column_widths = [15, 10, 20, 20, 30, 15, 10, 10, 10, 15, 15, 15, 20, 30, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 表头
    headers = ['字段名称', '工厂', 'BOM可选文本', '父项物料号', '物料名称', '生效日期', 
               'BOM用途', '可选BOM', 'BOM状态', '基本数量', '基本单位', '行项目号', 
               '项目类别', '子项物料号', '子项物料描述', '子项数量', '子项单位']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 按配方名称排序
    all_items.sort(key=lambda x: x['recipe_name'])
    
    # 数据行
    row = 2
    current_recipe = None
    bom_counter = 0
    
    for item in all_items:
        # 如果是新配方，更新计数器
        if current_recipe != item['recipe_name']:
            current_recipe = item['recipe_name']
            bom_counter += 1
        
        # 每行都填写完整的数据
        ws.cell(row=row, column=1, value='')  # 字段名称
        ws.cell(row=row, column=2, value='P060')  # 工厂
        ws.cell(row=row, column=3, value=item['recipe_description'])  # BOM可选文本
        ws.cell(row=row, column=4, value=data['parent_material_code'])  # 父项物料号
        ws.cell(row=row, column=5, value=data['parent_material_name'])  # 物料名称
        ws.cell(row=row, column=6, value='')  # 生效日期
        ws.cell(row=row, column=7, value='1')  # BOM用途
        ws.cell(row=row, column=8, value=f"{bom_counter:02d}")  # 可选BOM
        ws.cell(row=row, column=9, value='01')  # BOM状态
        ws.cell(row=row, column=10, value=data['basic_quantity'])  # 基本数量
        ws.cell(row=row, column=11, value=data['basic_unit'])  # 基本单位
        ws.cell(row=row, column=12, value=item['line_number'])  # 行项目号
        ws.cell(row=row, column=13, value=item['project_category'])  # 项目类别
        ws.cell(row=row, column=14, value=item['material_code'])  # 子项物料号
        ws.cell(row=row, column=15, value=item['material_name'])  # 子项物料描述
        ws.cell(row=row, column=16, value=item['quantity'] * data['basic_quantity'])  # 子项数量 = 配方数量 × 基本数量
        ws.cell(row=row, column=17, value=item['unit'])  # 子项单位
        row += 1
    
    # 保存文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return jsonify({
        'success': True,
        'message': 'BOM表生成成功',
        'file_path': temp_file.name
    })

@app.route('/api/download_bom/<path:file_path>')
def download_bom(file_path):
    try:
        return send_file(file_path, as_attachment=True, download_name='BOM表.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/recipe', methods=['POST'])
@login_required
def create_recipe():
    try:
        data = request.json
        
        # 检查配方名称是否已存在，如果存在则进行覆盖处理
        existing_recipe = Recipe.query.filter_by(recipe_name=data['name']).first()
        if existing_recipe:
            # 删除旧的配方项
            RecipeItem.query.filter_by(recipe_id=existing_recipe.id).delete()
            
            # 更新配方基本信息
            existing_recipe.description = data.get('description', '')
            existing_recipe.product_category = data.get('product_category', '')
            existing_recipe.updated_at = get_shanghai_time()
            
            # 添加新的配方项
            for item in data['items']:
                recipe_item = RecipeItem(
                    recipe_id=existing_recipe.id,
                    material_code=str(item['material_code']).strip(),
                    material_name=str(item['material_name']).strip(),
                    quantity=float(item['quantity']),
                    unit=str(item.get('unit', 'KG')).strip(),
                    line_number=str(item.get('line_number', '')).strip(),
                    project_category=str(item.get('project_category', 'L')).strip()
                )
                db.session.add(recipe_item)
            
            db.session.commit()
            return jsonify({
                'success': True, 
                'id': existing_recipe.id, 
                'message': f'配方 "{data["name"]}" 已存在，已覆盖原有内容'
            })
        
        # 验证配方项数据
        if not data.get('items') or len(data['items']) == 0:
            return jsonify({'success': False, 'message': '配方项不能为空'}), 400
        
        # 验证每个配方项的必填字段
        for i, item in enumerate(data['items']):
            if not item.get('material_code') or not item.get('material_name'):
                return jsonify({'success': False, 'message': f'第{i+1}个配方项：物料编码和物料名称不能为空'}), 400
            
            if item.get('quantity') is None or item.get('quantity') == '':
                return jsonify({'success': False, 'message': f'第{i+1}个配方项：数量不能为空'}), 400
            
            try:
                quantity = float(item['quantity'])
                if quantity <= 0:
                    return jsonify({'success': False, 'message': f'第{i+1}个配方项：数量必须大于0'}), 400
            except (ValueError, TypeError):
                return jsonify({'success': False, 'message': f'第{i+1}个配方项：数量必须是有效的数字'}), 400
        
        # 创建配方
        recipe = Recipe(
            recipe_name=data['name'],
            description=data.get('description', ''),
            product_category=data.get('product_category', '')
        )
        db.session.add(recipe)
        db.session.flush()  # 获取recipe.id
        
        # 添加配方项
        for item in data['items']:
            recipe_item = RecipeItem(
                recipe_id=recipe.id,
                material_code=str(item['material_code']).strip(),
                material_name=str(item['material_name']).strip(),
                quantity=float(item['quantity']),
                unit=str(item.get('unit', 'KG')).strip(),  # 默认KG
                line_number=str(item.get('line_number', '')).strip(),
                project_category=str(item.get('project_category', 'L')).strip()
            )
            db.session.add(recipe_item)
        
        db.session.commit()
        return jsonify({'success': True, 'id': recipe.id, 'message': '配方创建成功'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'创建配方失败：{str(e)}'}), 500

@app.route('/api/recipe/template')
@login_required
def download_recipe_template():
    """下载配方导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "配方导入模板"
    
    # 设置列宽 - A-I列分别为：配方名称、配方描述、行号、物料编码、物料名称、数量、单位、类别、产品类别
    column_widths = [20, 25, 12, 18, 25, 12, 12, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 表头 - 9列
    headers = ['配方名称', '配方描述', '行号', '物料编码', '物料名称', '数量', '单位', '类别', '产品类别']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 示例数据 - 包含配方名称、描述和产品类别，单位默认KG
    example_data = [
        ['示例配方1', '这是一个示例配方', '0010', 'M001', '原材料A', 2.5, 'KG', 'L', '电子产品'],
        ['示例配方1', '这是一个示例配方', '0020', 'M002', '原材料B', 5, 'KG', 'L', '电子产品'],
        ['示例配方2', '另一个示例配方', '0010', 'M003', '原材料C', 1.5, 'KG', 'L', '机械零件'],
        ['示例配方2', '另一个示例配方', '0020', 'M004', '原材料D', 1.0, 'KG', 'N', '机械零件']
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name='配方导入模板.xlsx')

@app.route('/api/recipe/import', methods=['POST'])
@login_required
def import_recipe():
    """导入配方"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # 从Excel文件中读取配方信息，而不是从表单
        # 模板格式：A-I列分别为：配方名称、配方描述、行号、物料编码、物料名称、数量、单位、类别、产品类别
        
        # 读取配方项并分组
        recipe_groups = {}  # 按配方名称分组
        
        for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
            recipe_name = ws.cell(row=row, column=1).value
            recipe_description = ws.cell(row=row, column=2).value
            line_number = ws.cell(row=row, column=3).value
            material_code = ws.cell(row=row, column=4).value
            material_name = ws.cell(row=row, column=5).value
            quantity = ws.cell(row=row, column=6).value
            unit = ws.cell(row=row, column=7).value
            project_category = ws.cell(row=row, column=8).value or 'L'
            product_category = ws.cell(row=row, column=9).value or ''
            
            # 检查必填字段
            if not all([recipe_name, line_number, material_code, material_name, quantity, unit]):
                continue  # 跳过不完整的行
            
            # 按配方名称分组
            if recipe_name not in recipe_groups:
                recipe_groups[recipe_name] = {
                    'description': recipe_description or '',
                    'product_category': product_category or '',
                    'items': []
                }
            
            recipe_groups[recipe_name]['items'].append({
                'line_number': str(line_number),
                'material_code': str(material_code),
                'material_name': str(material_name),
                'quantity': float(quantity),
                'unit': str(unit),
                'project_category': str(project_category)
            })
        
        # 创建配方和配方项
        total_items = 0
        created_recipes = []
        updated_recipes = []
        
        for recipe_name, recipe_data in recipe_groups.items():
            # 检查配方名称是否已存在
            existing_recipe = Recipe.query.filter_by(recipe_name=recipe_name).first()
            
            if existing_recipe:
                # 配方名称已存在，进行覆盖处理
                existing_recipe.description = recipe_data['description']
                existing_recipe.product_category = recipe_data['product_category']
                existing_recipe.updated_at = get_shanghai_time()
                
                # 删除旧的配方项
                RecipeItem.query.filter_by(recipe_id=existing_recipe.id).delete()
                
                # 添加新的配方项
                for item_data in recipe_data['items']:
                    recipe_item = RecipeItem(
                        recipe_id=existing_recipe.id,
                        material_code=item_data['material_code'],
                        material_name=item_data['material_name'],
                        quantity=item_data['quantity'],
                        unit=item_data['unit'],
                        line_number=item_data['line_number'],
                        project_category=item_data['project_category']
                    )
                    db.session.add(recipe_item)
                    total_items += 1
                
                updated_recipes.append(recipe_name)
            else:
                # 创建新配方
                recipe = Recipe(
                    recipe_name=recipe_name,
                    description=recipe_data['description'],
                    product_category=recipe_data['product_category']
                )
                db.session.add(recipe)
                db.session.flush()  # 获取recipe.id
                
                # 创建配方项
                for item_data in recipe_data['items']:
                    recipe_item = RecipeItem(
                        recipe_id=recipe.id,
                        material_code=item_data['material_code'],
                        material_name=item_data['material_name'],
                        quantity=item_data['quantity'],
                        unit=item_data['unit'],
                        line_number=item_data['line_number'],
                        project_category=item_data['project_category']
                    )
                    db.session.add(recipe_item)
                    total_items += 1
                
                created_recipes.append(recipe_name)
        
        db.session.commit()
        
        # 构建返回消息
        message_parts = []
        if created_recipes:
            message_parts.append(f'新建 {len(created_recipes)} 个配方')
        if updated_recipes:
            message_parts.append(f'覆盖 {len(updated_recipes)} 个配方')
        
        message = f'配方导入成功，{", ".join(message_parts)}，共 {total_items} 个配方项'
        
        return jsonify({
            'success': True, 
            'message': message,
            'recipe_count': len(created_recipes) + len(updated_recipes),
            'item_count': total_items,
            'created_recipes': created_recipes,
            'updated_recipes': updated_recipes
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'}), 400

@app.route('/api/recipe/export/<int:recipe_id>')
@login_required
def export_recipe(recipe_id):
    """导出配方到Excel"""
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({'success': False, 'message': '配方不存在'}), 404
    
    items = RecipeItem.query.filter_by(recipe_id=recipe_id).order_by(RecipeItem.line_number).all()
    
    # 创建Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    # 过滤sheet名称中的无效字符
    safe_title = "".join(c for c in recipe.recipe_name if c not in r'\/:*?"<>|')
    ws.title = safe_title[:31] if len(safe_title) > 31 else safe_title  # Excel sheet名称最大31字符
    
    # 设置列宽
    column_widths = [15, 20, 30, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 配方信息
    ws.cell(row=1, column=1, value='配方名称：')
    ws.cell(row=1, column=2, value=recipe.recipe_name)
    ws.cell(row=2, column=1, value='配方描述：')
    ws.cell(row=2, column=2, value=recipe.description or '')
    
    # 表头
    headers = ['行号', '物料编码', '物料名称', '数量', '单位', '类别']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 配方项数据
    for row_idx, item in enumerate(items, 5):
        ws.cell(row=row_idx, column=1, value=item.line_number)
        ws.cell(row=row_idx, column=2, value=item.material_code)
        ws.cell(row=row_idx, column=3, value=item.material_name)
        ws.cell(row=row_idx, column=4, value=item.quantity)
        ws.cell(row=row_idx, column=5, value=item.unit)
        ws.cell(row=row_idx, column=6, value=item.project_category)
    
    # 保存文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name=f'{recipe.recipe_name}.xlsx')

@app.route('/api/recipe/export_all')
@login_required
def export_all_recipes():
    """批量导出所有配方到单个表格"""
    try:
        # 获取所有活跃的配方
        recipes = Recipe.query.filter_by(is_active=True).order_by(Recipe.recipe_name).all()
        
        if not recipes:
            return jsonify({'error': '没有找到可导出的配方'}), 404
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "所有配方"
        
        # 设置列宽
        column_widths = [20, 40, 20, 15, 20, 30, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 表头 A-I列：配方名称、配方描述、行号、物料编码、物料名称、数量、单位、类别、产品类别
        headers = ['配方名称', '配方描述', '行号', '物料编码', '物料名称', '数量', '单位', '类别', '产品类别']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        row = 2
        for recipe in recipes:
            # 获取配方项
            items = RecipeItem.query.filter_by(recipe_id=recipe.id).order_by(RecipeItem.line_number).all()
            
            # 为每个配方项创建一行数据
            for item in items:
                ws.cell(row=row, column=1, value=recipe.recipe_name)  # A列：配方名称
                ws.cell(row=row, column=2, value=recipe.description or '')  # B列：配方描述
                ws.cell(row=row, column=3, value=item.line_number)  # C列：行号
                ws.cell(row=row, column=4, value=item.material_code)  # D列：物料编码
                ws.cell(row=row, column=5, value=item.material_name)  # E列：物料名称
                ws.cell(row=row, column=6, value=item.quantity)  # F列：数量
                ws.cell(row=row, column=7, value=item.unit)  # G列：单位
                ws.cell(row=row, column=8, value=item.project_category)  # H列：类别
                ws.cell(row=row, column=9, value=recipe.product_category or '')  # I列：产品类别
                row += 1
        
        # 保存文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(temp_file.name, as_attachment=True, download_name='所有配方.xlsx')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400





@app.route('/api/download_batch_bom/<path:file_path>')
def download_batch_bom(file_path):
    """下载批量生成的BOM文件"""
    try:
        # 修复文件路径问题
        if file_path.startswith('C:'):
            # 处理Windows路径
            file_path = file_path.replace('C:', 'C:\\')
        elif file_path.startswith('C%3A'):
            # 处理URL编码的路径
            file_path = file_path.replace('C%3A', 'C:\\')
        
        # 确保文件存在
        if not os.path.exists(file_path):
            return jsonify({'error': '文件不存在'}), 404
            
        return send_file(file_path, as_attachment=True, download_name='批量BOM表.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/bom/batch_generate_table', methods=['POST'])
def batch_generate_bom_from_table():
    """从弹窗表格数据批量生成BOM"""
    try:
        data = request.json
        bom_data = data.get('bom_data', [])
        
        if not bom_data:
            return jsonify({'success': False, 'message': '没有提供BOM数据'}), 400
        
        bom_items = []
        errors = []
        
        for item in bom_data:
            # 检查必填字段
            if not all([item.get('parent_material_code'), item.get('parent_material_name'), 
                       item.get('basic_quantity'), item.get('basic_unit'), item.get('recipe_names')]):
                errors.append(f"第{item.get('line_number', '?')}行：必填字段不完整")
                continue
            
            # 验证数量
            try:
                basic_quantity = float(item['basic_quantity'])
                if basic_quantity <= 0:
                    errors.append(f"第{item.get('line_number', '?')}行：基本数量必须大于0")
                    continue
            except:
                errors.append(f"第{item.get('line_number', '?')}行：基本数量必须是数字")
                continue
            
            # 查找配方ID
            recipe_ids = []
            for recipe_name in item['recipe_names']:
                # 确保配方名称被正确清理（去除前后空格）
                recipe_name = recipe_name.strip()
                if not recipe_name:  # 跳过空的配方名称
                    continue
                    
                recipe = Recipe.query.filter_by(recipe_name=recipe_name, is_active=True).first()
                if recipe:
                    recipe_ids.append(recipe.id)
                else:
                    errors.append(f"第{item.get('line_number', '?')}行：配方'{recipe_name}'不存在")
                    break
            
            if len(recipe_ids) != len(item['recipe_names']):
                continue
            
            # 添加到BOM项目列表
            bom_items.append({
                'parent_material_code': str(item['parent_material_code']),
                'parent_material_name': str(item['parent_material_name']),
                'basic_quantity': basic_quantity,
                'basic_unit': str(item['basic_unit']),
                'recipe_ids': recipe_ids,
                'remark': item.get('notes', '')
            })
        
        if errors:
            return jsonify({
                'success': False, 
                'message': f'生成失败，发现 {len(errors)} 个错误',
                'errors': errors
            }), 400
        
        if not bom_items:
            return jsonify({'success': False, 'message': '未找到有效的BOM数据'}), 400
        
        # 生成所有BOM的Excel文件
        wb_result = openpyxl.Workbook()
        ws_result = wb_result.active
        ws_result.title = "批量BOM表"
        
        # 确保只有一个工作表
        while len(wb_result.sheetnames) > 1:
            wb_result.remove(wb_result.sheetnames[-1])
        
        print(f"DEBUG: 创建工作簿，工作表数量: {len(wb_result.sheetnames)}")
        print(f"DEBUG: 工作表名称: {wb_result.sheetnames}")
        
        # 设置列宽 (A-Q列)
        column_widths = [15, 10, 20, 20, 30, 15, 10, 10, 10, 15, 15, 15, 20, 30, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws_result.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 表头 (A-Q列)
        headers = ['字段名称', '工厂', 'BOM可选文本', '父项物料号', '物料名称', '生效日期', 
                   'BOM用途', '可选BOM', 'BOM状态', '基本数量', '基本单位', '行项目号', 
                   '项目类别', '子项物料号', '子项物料描述', '子项数量', '子项单位']
        
        for col, header in enumerate(headers, 1):
            cell = ws_result.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        row = 2
        
        for bom_item in bom_items:
            # 为每个父物料重新开始计数
            bom_counter = 0
            
            # 获取所有选中的配方信息
            all_items = []
            
            for recipe_id in bom_item['recipe_ids']:
                recipe = Recipe.query.get(recipe_id)
                if recipe:
                    items = RecipeItem.query.filter_by(recipe_id=recipe_id).order_by(RecipeItem.line_number).all()
                    # 为每个配方项添加配方信息
                    for item in items:
                        all_items.append({
                            'recipe_name': recipe.recipe_name,
                            'recipe_description': recipe.description,
                            'line_number': item.line_number,
                            'material_code': item.material_code,
                            'material_name': item.material_name,
                            'quantity': item.quantity,
                            'unit': item.unit,
                            'project_category': item.project_category
                        })
            
            # 按配方名称排序
            all_items.sort(key=lambda x: x['recipe_name'])
            
            # 为每个BOM项生成数据行
            current_recipe = None
            
            for item in all_items:
                # 如果是新配方，更新计数器
                if current_recipe != item['recipe_name']:
                    current_recipe = item['recipe_name']
                    bom_counter += 1
                
                # 每行都填写完整的数据
                ws_result.cell(row=row, column=1, value='')  # 字段名称
                ws_result.cell(row=row, column=2, value='P060')  # 工厂
                ws_result.cell(row=row, column=3, value=item['recipe_description'])  # BOM可选文本
                ws_result.cell(row=row, column=4, value=bom_item['parent_material_code'])  # 父项物料号
                ws_result.cell(row=row, column=5, value=bom_item['parent_material_name'])  # 物料名称
                ws_result.cell(row=row, column=6, value='')  # 生效日期
                ws_result.cell(row=row, column=7, value='1')  # BOM用途
                ws_result.cell(row=row, column=8, value=f"{bom_counter:02d}")  # 可选BOM
                ws_result.cell(row=row, column=9, value='01')  # BOM状态
                ws_result.cell(row=row, column=10, value=bom_item['basic_quantity'])  # 基本数量
                ws_result.cell(row=row, column=11, value=bom_item['basic_unit'])  # 基本单位
                ws_result.cell(row=row, column=12, value=item['line_number'])  # 行项目号
                ws_result.cell(row=row, column=13, value=item['project_category'])  # 项目类别
                ws_result.cell(row=row, column=14, value=item['material_code'])  # 子项物料号
                ws_result.cell(row=row, column=15, value=item['material_name'])  # 子项物料描述
                ws_result.cell(row=row, column=16, value=item['quantity'] * bom_item['basic_quantity'])  # 子项数量 = 配方数量 × 基本数量
                ws_result.cell(row=row, column=17, value=item['unit'])  # 子项单位
                
                row += 1
        
        # 保存文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb_result.save(temp_file.name)
        temp_file.close()
        
        return jsonify({
            'success': True,
            'message': f'成功生成 {len(bom_items)} 个BOM表！',
            'file_path': temp_file.name,
            'bom_count': len(bom_items)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'批量生成失败：{str(e)}'}), 400

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # 初始化一些示例数据
        if not Recipe.query.first():
            recipe = Recipe(recipe_name='标准配方1', description='标准产品配方')
            db.session.add(recipe)
            db.session.commit()
            
            items = [
                RecipeItem(recipe_id=recipe.id, material_code='M001', material_name='原材料A', 
                          quantity=2.5, unit='KG', line_number='0010', project_category='L'),
                RecipeItem(recipe_id=recipe.id, material_code='M002', material_name='原材料B', 
                          quantity=5, unit='PCS', line_number='0020', project_category='L'),
            ]
            for item in items:
                db.session.add(item)
            
            db.session.commit()
    
    # 生产环境配置
    import os
    
    # 检查环境变量，决定使用哪种服务器
    if os.environ.get('FLASK_ENV') == 'production' or os.environ.get('USE_PRODUCTION_SERVER'):
        # 使用生产级WSGI服务器
        try:
            from waitress import serve
            print("🚀 使用生产级Waitress WSGI服务器启动...")
            print("📍 访问地址：")
            print("   本机访问：http://localhost:5000")
            print("   局域网访问：http://0.0.0.0:5000")
            print("✅ 生产环境模式，无开发服务器警告")
            serve(app, host='0.0.0.0', port=5000, threads=4)
        except ImportError:
            print("⚠️  Waitress未安装，回退到Flask内置服务器")
            print("💡 建议安装：pip install waitress")
            app.run(
                host='0.0.0.0',  # 允许局域网访问
                port=5000, 
                debug=False,  # 关闭调试模式，提高安全性
                threaded=True,  # 启用多线程，支持并发访问
                processes=1  # 单进程模式，避免数据库连接问题
            )
    else:
        # 开发环境配置
        print("🔧 使用Flask内置服务器启动（开发模式）")
        print("⚠️  注意：这是开发服务器，不适合生产环境")
        app.run(
            host='0.0.0.0',  # 允许局域网访问
            port=5000, 
            debug=False,  # 关闭调试模式，提高安全性
            threaded=True,  # 启用多线程，支持并发访问
            processes=1  # 单进程模式，避免数据库连接问题
        )
